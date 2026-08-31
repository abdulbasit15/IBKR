"""ZTH Trade Journal — turn a Tradovate order export into ZTH tracker rows.

Deterministic, dependency-light (csv from stdlib; openpyxl only needed to write
into the workbook). This is the *source of truth* for how bracket orders are
grouped and scored, so the journaling is reproducible on any machine and does
not rely on an AI re-deriving the logic each time.

Usage
-----
    # preview only (no file changes):
    python zth_journal.py "tradovate-orders-all-....csv" --dry-run

    # write into the workbook (a timestamped backup is always made first):
    python zth_journal.py "tradovate-orders-all-....csv"

    # options:
    #   --workbook "ZTH Trade Tracker - AB.xlsx"   (default: next to this script)
    #   --sheet    "Trade Tracker - Eval"          (default target sheet)
    #   --dry-run                                  (preview, do not write)

Conventions (agreed with the trader)
------------------------------------
* A *bracket* = one filled entry (Type=Limit) + a Take Profit leg + a Stop Loss
  leg. The leg that actually *filled* is the exit.
* Entry price  = entry's Avg Fill Price.
* Exit price   = filled exit leg's Avg Fill Price (ACTUAL fill, with decimals).
* Stop Loss    = Stop Loss leg's Stop Price (the level you set).
* Take Profit  = Take Profit leg's Limit Price (the level you set).
* Direction    = Buy entry -> Long, Sell entry -> Short. Size = entry Filled Qty.
* Asset root   = futures symbol with the month+year code stripped (MNQU6 -> MNQ).
* Only the INPUT columns are written (Date, Time, Asset, Direction, Entry, Exit,
  Size, Stop Loss, Take Profit, Instrument Level). Formula columns (RR, P&L,
  Cumulative...) and the ZTH number column are never touched.
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import io
import os
import re
import shutil
import sys

# --- domain constants --------------------------------------------------------

EXIT_TYPES = {"Take Profit", "Stop Loss"}
MONTH_CODES = "FGHJKMNQUVXZ"  # futures delivery month letters

# Point multipliers ($ per 1.0 price move per contract). Mirrors the workbook's
# Multipliers sheet; used for the printed preview and tests. The workbook's own
# formulas remain the authority once written.
DEFAULT_MULTIPLIERS = {
    "ES": 50, "MES": 5, "NQ": 20, "MNQ": 2, "YM": 5, "RTY": 50,
    "GC": 100, "MGC": 10, "CL": 1000, "MCL": 100, "QM": 500,
}

_UPDATE_TIME_FORMATS = ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M")


def root_symbol(symbol: str) -> str:
    """Strip a futures month+year code: 'MNQU6' -> 'MNQ', 'ESH7' -> 'ES'.

    If the tail does not look like <month-letter><1-2 digits>, return as-is.
    """
    symbol = (symbol or "").strip()
    m = re.match(rf"^([A-Z]+?)[{MONTH_CODES}]\d{{1,2}}$", symbol)
    return m.group(1) if m else symbol


def _to_float(value):
    value = (value or "").strip()
    if value == "":
        return None
    try:
        return round(float(value), 4)
    except ValueError:
        return None


def _to_int(value):
    value = (value or "").strip()
    if value == "":
        return 0
    try:
        return int(float(value))
    except ValueError:
        return 0


def _parse_dt(value):
    value = (value or "").strip()
    for fmt in _UPDATE_TIME_FORMATS:
        try:
            return _dt.datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


# --- parsing -----------------------------------------------------------------

def parse_csv_text(text: str) -> list[dict]:
    """Parse the raw CSV text into a list of normalized order dicts."""
    reader = csv.DictReader(io.StringIO(text))
    orders = []
    for row in reader:
        row = {(k or "").strip(): v for k, v in row.items()}
        symbol = (row.get("Symbol") or "").strip()
        if not symbol:
            continue
        orders.append({
            "symbol": symbol,
            "side": (row.get("Side") or "").strip(),
            "type": (row.get("Type") or "").strip(),
            "status": (row.get("Status") or "").strip(),
            "qty": _to_int(row.get("Qty")),
            "filled_qty": _to_int(row.get("Filled Qty")),
            "limit_price": _to_float(row.get("Limit Price")),
            "stop_price": _to_float(row.get("Stop Price")),
            "avg_fill": _to_float(row.get("Avg Fill Price")),
            "update_dt": _parse_dt(row.get("Update Time")),
            "update_raw": (row.get("Update Time") or "").strip(),
            "order_id": (row.get("Order ID") or "").strip(),
        })
    return orders


def parse_orders(path: str) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig") as f:
        return parse_csv_text(f.read())


# --- grouping ----------------------------------------------------------------

def group_into_trades(orders: list[dict]) -> list[dict]:
    """Group orders into bracket trades.

    Walk each symbol's orders in time order. A filled non-exit order opens a
    bracket; the following Take Profit / Stop Loss legs attach to it. The first
    filled leg is the exit. A bracket closes when both legs are seen (or the
    next entry for that symbol arrives).
    """
    ordered = sorted(
        orders,
        key=lambda o: (o["update_dt"] or _dt.datetime.max, o.get("order_id", "")),
    )
    trades: list[dict] = []
    open_bracket: dict[str, dict] = {}

    def _flush(sym):
        b = open_bracket.pop(sym, None)
        if b and b.get("exit_leg"):
            trades.append(b)

    for o in ordered:
        sym = o["symbol"]
        if o["type"] in EXIT_TYPES:
            b = open_bracket.get(sym)
            if b is None:
                continue  # orphan exit leg with no matching entry
            if o["type"] == "Take Profit":
                b["tp_leg"] = o
            else:
                b["sl_leg"] = o
            if o["status"] == "Filled" and b.get("exit_leg") is None:
                b["exit_leg"] = o
            if b.get("tp_leg") and b.get("sl_leg"):
                _flush(sym)
        else:
            if o["status"] != "Filled":
                continue  # unfilled entry -> no trade
            _flush(sym)  # close any prior open bracket for this symbol
            open_bracket[sym] = {"entry": o, "symbol": sym}

    for sym in list(open_bracket):
        _flush(sym)

    trades.sort(key=lambda t: (t["entry"]["update_dt"] or _dt.datetime.max))
    return trades


# --- scoring -----------------------------------------------------------------

def compute_journal_row(trade: dict, multipliers: dict | None = None) -> dict:
    multipliers = multipliers or DEFAULT_MULTIPLIERS
    entry = trade["entry"]
    exit_leg = trade["exit_leg"]
    tp_leg = trade.get("tp_leg")
    sl_leg = trade.get("sl_leg")

    asset = root_symbol(entry["symbol"])
    direction = "Long" if entry["side"] == "Buy" else "Short"
    entry_price = entry["avg_fill"]
    exit_price = exit_leg["avg_fill"]
    size = entry["filled_qty"] or entry["qty"]
    sl_level = sl_leg["stop_price"] if sl_leg else None
    tp_level = tp_leg["limit_price"] if tp_leg else None

    mult = multipliers.get(asset)
    pnl = None
    if entry_price is not None and exit_price is not None and mult is not None:
        move = (exit_price - entry_price) if direction == "Long" else (entry_price - exit_price)
        pnl = round(move * size * mult, 2)

    win_loss = ""
    if pnl is not None:
        win_loss = "Win" if pnl > 0 else ("Loss" if pnl < 0 else "")

    dt = entry["update_dt"]
    return {
        "update_dt": dt,
        "date": dt.date() if dt else None,
        "time": dt.time().replace(second=0) if dt else None,
        "asset": asset,
        "direction": direction,
        "entry": entry_price,
        "exit": exit_price,
        "size": size,
        "stop_loss": sl_level,
        "take_profit": tp_level,
        "instrument_level": asset,
        "pnl": pnl,
        "win_loss": win_loss,
        "multiplier_found": mult is not None,
    }


# --- workbook writing --------------------------------------------------------

# header text -> journal-row key (input columns only; formulas untouched)
_COLUMN_MAP = {
    "Date": "date",
    "Time": "time",
    "Asset Traded": "asset",
    "Direction": "direction",
    "Entry Price": "entry",
    "Exit Price": "exit",
    "Position Size": "size",
    "Stop Loss": "stop_loss",
    "Take Profit": "take_profit",
    # header is misspelled "Intrument Level" in the workbook; match both.
    "Intrument Level": "instrument_level",
    "Instrument Level": "instrument_level",
}


def _header_columns(ws):
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str):
            cols[v.strip()] = c
    return cols


def write_rows_to_workbook(workbook_path, sheet_name, rows, dry_run=False):
    """Append journal rows into the workbook sheet. Returns a result dict.

    Always creates a timestamped backup before saving. Only input columns are
    written; formula columns and the ZTH number column are left untouched.
    Duplicate trades (same asset/date/time/entry/exit) are skipped.
    """
    import openpyxl  # imported lazily so parsing/tests need no third-party dep

    result = {"written": [], "skipped_dupes": [], "warnings": [], "backup": None}
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Sheets: {wb.sheetnames}")
    ws = wb[sheet_name]
    cols = _header_columns(ws)

    entry_col = cols.get("Entry Price")
    if not entry_col:
        raise ValueError("Could not find an 'Entry Price' column in the sheet header.")

    # existing trades (for dedupe) + last formula row
    existing = set()
    last_formula_row = 1
    for r in range(2, ws.max_row + 1):
        e = ws.cell(r, entry_col).value
        if isinstance(e, (int, float)):
            key = (
                ws.cell(r, cols.get("Asset Traded", 4)).value,
                ws.cell(r, cols.get("Date", 2)).value.date()
                if isinstance(ws.cell(r, cols.get("Date", 2)).value, _dt.datetime) else None,
                ws.cell(r, cols.get("Time", 3)).value,
                ws.cell(r, entry_col).value,
                ws.cell(r, cols.get("Exit Price", 7)).value,
            )
            existing.add(key)
        pnl_col = cols.get("P&L Amount")
        if pnl_col and isinstance(ws.cell(r, pnl_col).value, str) and ws.cell(r, pnl_col).value.startswith("="):
            last_formula_row = r

    # number formats copied from an existing dated row so Date/Time render right
    date_fmt = time_fmt = None
    for r in range(2, ws.max_row + 1):
        dv = ws.cell(r, cols.get("Date", 2)).value
        if isinstance(dv, _dt.datetime):
            date_fmt = ws.cell(r, cols["Date"]).number_format
            time_fmt = ws.cell(r, cols["Time"]).number_format
            break

    # first empty data row = first row (>=2) whose Entry Price is blank
    next_row = 2
    while ws.cell(next_row, entry_col).value not in (None, ""):
        next_row += 1

    for row in rows:
        dupe_key = (row["asset"],
                    row["date"] if not isinstance(row["date"], _dt.datetime) else row["date"].date(),
                    row["time"], row["entry"], row["exit"])
        if dupe_key in existing:
            result["skipped_dupes"].append(row)
            continue
        if not row["multiplier_found"]:
            result["warnings"].append(
                f"Asset {row['asset']!r} not in Multipliers sheet — P&L will error until added.")
        if next_row > last_formula_row:
            result["warnings"].append(
                f"Row {next_row} is past the last formula row ({last_formula_row}); "
                f"RR/P&L formulas may be missing there — extend them in Excel.")

        for header, key in _COLUMN_MAP.items():
            if header not in cols:
                continue
            c = cols[header]
            val = row[key]
            if val is None:
                continue
            if key == "date":
                cell = ws.cell(next_row, c, _dt.datetime.combine(val, _dt.time()) if not isinstance(val, _dt.datetime) else val)
                if date_fmt:
                    cell.number_format = date_fmt
            elif key == "time":
                cell = ws.cell(next_row, c, val)
                if time_fmt:
                    cell.number_format = time_fmt
            else:
                ws.cell(next_row, c, val)

        existing.add(dupe_key)
        result["written"].append((next_row, row))
        next_row += 1

    if dry_run or not result["written"]:
        result["dry_run"] = True
        return result

    stamp = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    base, ext = os.path.splitext(workbook_path)
    backup = f"{base}.backup-{stamp}{ext}"
    shutil.copy2(workbook_path, backup)
    result["backup"] = backup
    wb.save(workbook_path)
    return result


# --- CLI ---------------------------------------------------------------------

def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def print_preview(rows):
    headers = ["Date", "Time", "Asset", "Dir", "Entry", "Exit", "Size", "SL", "TP", "P&L", "W/L"]
    print("  " + " | ".join(f"{h:>9}" for h in headers))
    net = 0.0
    for r in rows:
        if r["pnl"] is not None:
            net += r["pnl"]
        cells = [
            r["date"].strftime("%m/%d/%Y") if r["date"] else "",
            r["time"].strftime("%H:%M") if r["time"] else "",
            r["asset"], r["direction"], _fmt(r["entry"]), _fmt(r["exit"]),
            _fmt(r["size"]), _fmt(r["stop_loss"]), _fmt(r["take_profit"]),
            _fmt(r["pnl"]), r["win_loss"],
        ]
        print("  " + " | ".join(f"{str(c):>9}" for c in cells))
    print(f"\n  Trades: {len(rows)}   Net P&L: {net:+.2f}")


def main(argv=None):
    p = argparse.ArgumentParser(description="Journal Tradovate bracket orders into the ZTH tracker.")
    p.add_argument("csv", help="Path to the Tradovate orders CSV export.")
    here = os.path.dirname(os.path.abspath(__file__))
    p.add_argument("--workbook", default=os.path.join(here, "ZTH Trade Tracker - AB.xlsx"))
    p.add_argument("--sheet", default="Trade Tracker - Eval")
    p.add_argument("--dry-run", action="store_true", help="Preview only; do not write.")
    args = p.parse_args(argv)

    orders = parse_orders(args.csv)
    trades = group_into_trades(orders)
    rows = [compute_journal_row(t) for t in trades]
    rows.sort(key=lambda r: (r["update_dt"] or _dt.datetime.max))

    print(f"Parsed {len(orders)} orders -> {len(trades)} bracket trades from:\n  {args.csv}\n")
    print_preview(rows)

    if args.dry_run:
        print("\n[dry-run] No changes written.")
        return 0

    res = write_rows_to_workbook(args.workbook, args.sheet, rows, dry_run=False)
    print()
    for w in res["warnings"]:
        print(f"  WARNING: {w}")
    for row in res["skipped_dupes"]:
        print(f"  skipped duplicate: {row['asset']} {row['date']} entry={row['entry']}")
    if res.get("backup"):
        print(f"  backup: {res['backup']}")
    print(f"  wrote {len(res['written'])} trade(s) to '{args.sheet}' in:\n  {args.workbook}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
