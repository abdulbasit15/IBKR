"""Read trades from and write analysis back to the ZTH Trade Tracker workbook."""
from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config

# 1-indexed column positions in the "Trade Tracker - Eval" sheet.
COL = {
    "zth": 1, "date": 2, "time": 3, "asset": 4, "direction": 5,
    "entry": 6, "exit": 7, "size": 8, "stop": 9, "target": 10,
    "rr_targeted": 11, "rr_realized": 12, "winloss": 13,
    "pnl": 14, "pnl_pct": 15, "cum_pnl": 16,
    "setup": 17, "notes": 18, "instrument_level": 19,
    "level_screenshot": 20, "trade_screenshot": 21,
}
FIRST_DATA_ROW = 2


@dataclass
class Trade:
    zth: object = None
    date: object = None
    time: object = None
    asset: str = ""
    direction: str = ""
    entry: float = None
    exit: float = None
    size: float = None
    stop: float = None
    target: float = None
    rr_targeted: float = None
    rr_realized: float = None
    winloss: str = ""
    pnl: float = None
    pnl_pct: float = None
    cum_pnl: float = None
    setup: str = ""
    notes: str = ""
    instrument_level: str = ""
    level_screenshot: str = ""
    trade_screenshot: str = ""

    @property
    def timestamp(self) -> datetime | None:
        if self.date is None:
            return None
        d = self.date.date() if isinstance(self.date, datetime) else self.date
        t = self.time if isinstance(self.time, time) else time(0, 0)
        return datetime.combine(d, t)

    def dedupe_key(self) -> tuple:
        d = self.date.date() if isinstance(self.date, datetime) else self.date
        t = self.time.strftime("%H:%M") if isinstance(self.time, time) else ""
        return (str(d), t, (self.asset or "").upper(), (self.direction or "").title(),
                round(float(self.entry), 4) if self.entry is not None else None)


def _cell(ws, row, key):
    return ws.cell(row=row, column=COL[key]).value


def read_trades(workbook: Path | None = None) -> list[Trade]:
    """Return the real (non-placeholder) trades using cached computed values."""
    wb_path = Path(workbook) if workbook else config.WORKBOOK
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb[config.TRADE_SHEET]
    trades: list[Trade] = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d, asset, entry = _cell(ws, r, "date"), _cell(ws, r, "asset"), _cell(ws, r, "entry")
        if d is None or not asset or entry is None:
            continue  # placeholder / blank row
        trades.append(Trade(
            zth=_cell(ws, r, "zth"), date=d, time=_cell(ws, r, "time"),
            asset=str(asset).strip().upper(),
            direction=str(_cell(ws, r, "direction") or "").strip().title(),
            entry=_num(entry), exit=_num(_cell(ws, r, "exit")),
            size=_num(_cell(ws, r, "size")), stop=_num(_cell(ws, r, "stop")),
            target=_num(_cell(ws, r, "target")),
            rr_targeted=_num(_cell(ws, r, "rr_targeted")),
            rr_realized=_num(_cell(ws, r, "rr_realized")),
            winloss=str(_cell(ws, r, "winloss") or "").strip(),
            pnl=_num(_cell(ws, r, "pnl")), pnl_pct=_num(_cell(ws, r, "pnl_pct")),
            cum_pnl=_num(_cell(ws, r, "cum_pnl")),
            setup=str(_cell(ws, r, "setup") or "").strip(),
            notes=str(_cell(ws, r, "notes") or "").strip(),
            instrument_level=str(_cell(ws, r, "instrument_level") or "").strip(),
        ))
    wb.close()
    return trades


def _num(v):
    if v is None or v == "" or isinstance(v, str):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def backup(workbook: Path | None = None) -> Path:
    wb_path = Path(workbook) if workbook else config.WORKBOOK
    config.BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = config.BACKUP_DIR / f"{wb_path.stem}.{stamp}.bak.xlsx"
    shutil.copy2(wb_path, dest)
    return dest


def write_setups(labels: dict, workbook: Path | None = None,
                 only_blank: bool = True) -> int:
    """Write setup labels into column Q for real trade rows (in read order).

    `labels` is keyed by the 0-based index of each real trade row, matching the
    order returned by read_trades(). Returns the number of cells written.
    """
    wb_path = Path(workbook) if workbook else config.WORKBOOK
    backup(wb_path)
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws = wb[config.TRADE_SHEET]
    i, written = 0, 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if _cell(ws, r, "date") is None or not _cell(ws, r, "asset") \
           or _cell(ws, r, "entry") is None:
            continue
        label = labels.get(i)
        cur = ws.cell(row=r, column=COL["setup"]).value
        if label and (not only_blank or not cur):
            ws.cell(row=r, column=COL["setup"], value=label)
            written += 1
        i += 1
    wb.save(wb_path)
    wb.close()
    return written


def ensure_multipliers(ws_mult) -> None:
    """Make sure every configured root has a row on the Multipliers sheet."""
    existing = {}
    for r in range(2, ws_mult.max_row + 1):
        name = ws_mult.cell(row=r, column=1).value
        if name:
            existing[str(name).strip().upper()] = r
    next_row = ws_mult.max_row + 1
    for root, mult in config.MULTIPLIERS.items():
        if root not in existing:
            ws_mult.cell(row=next_row, column=1, value=root)
            ws_mult.cell(row=next_row, column=2, value=mult)
            next_row += 1


def _row_formulas(r: int) -> dict[str, str]:
    """The workbook's per-row formulas, with the Cumulative P&L bug fixed."""
    return {
        "rr_targeted": f'=IF(AND(I{r}<>"",J{r}<>"",F{r}<>""), (J{r}-F{r})/(F{r}-I{r}), "")',
        "rr_realized": f'=IF(AND(F{r}<>"",G{r}<>"",I{r}<>""), (G{r}-F{r})/(F{r}-I{r}), "")',
        "winloss": f'=IF(N{r}>0,"Win",IF(N{r}<0,"Loss",""))',
        "pnl": (f'=IF(AND(F{r}<>"",G{r}<>"",H{r}<>"",D{r}<>""), '
                f'IF(E{r}="Long", (G{r}-F{r})*H{r}*VLOOKUP(D{r},Multipliers!A:B,2,FALSE), '
                f'(F{r}-G{r})*H{r}*VLOOKUP(D{r},Multipliers!A:B,2,FALSE)), "")'),
        "pnl_pct": (f'=IF(AND(N{r}<>"",F{r}<>"",H{r}<>"",D{r}<>""), '
                    f'N{r}/(F{r}*H{r}*VLOOKUP(D{r},Multipliers!A:B,2,FALSE)), "")'),
        # Fixed: previous row reference instead of the broken #REF!.
        "cum_pnl": (f"=N{r}" if r == FIRST_DATA_ROW else f"=P{r-1}+N{r}"),
    }


def _last_real_row(ws) -> int:
    last = FIRST_DATA_ROW - 1
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if _cell(ws, r, "date") is not None and _cell(ws, r, "asset"):
            last = r
    return last


def append_trades(new_trades: list[Trade], workbook: Path | None = None) -> int:
    """Append new trades, writing raw inputs + the workbook's formulas.

    Also repairs the Cumulative P&L formula across all existing data rows.
    Returns the number of rows written.
    """
    if not new_trades:
        return 0
    wb_path = Path(workbook) if workbook else config.WORKBOOK
    backup(wb_path)
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws = wb[config.TRADE_SHEET]
    if config.MULTIPLIER_SHEET in wb.sheetnames:
        ensure_multipliers(wb[config.MULTIPLIER_SHEET])

    last = _last_real_row(ws)
    last_zth = _cell(ws, last, "zth") if last >= FIRST_DATA_ROW else None
    try:
        next_zth = int(last_zth) + 1
    except (TypeError, ValueError):
        next_zth = (last - FIRST_DATA_ROW + 2)

    r = last + 1
    for t in new_trades:
        ws.cell(row=r, column=COL["zth"], value=next_zth)
        d = t.date.date() if isinstance(t.date, datetime) else t.date
        c = ws.cell(row=r, column=COL["date"], value=datetime.combine(d, time(0, 0)) if d else None)
        c.number_format = "m/d/yyyy"
        c = ws.cell(row=r, column=COL["time"], value=t.time)
        c.number_format = "h:mm"
        ws.cell(row=r, column=COL["asset"], value=t.asset)
        ws.cell(row=r, column=COL["direction"], value=t.direction)
        ws.cell(row=r, column=COL["entry"], value=t.entry)
        ws.cell(row=r, column=COL["exit"], value=t.exit)
        ws.cell(row=r, column=COL["size"], value=t.size)
        ws.cell(row=r, column=COL["stop"], value=t.stop)
        ws.cell(row=r, column=COL["target"], value=t.target)
        ws.cell(row=r, column=COL["setup"], value=t.setup or "")
        ws.cell(row=r, column=COL["notes"], value=t.notes or "")
        ws.cell(row=r, column=COL["instrument_level"], value=t.asset)
        for key, f in _row_formulas(r).items():
            ws.cell(row=r, column=COL[key], value=f)
        next_zth += 1
        r += 1

    # Repair cumulative formulas for every real row (fixes historic #REF!).
    for rr in range(FIRST_DATA_ROW, r):
        if _cell(ws, rr, "date") is not None:
            ws.cell(row=rr, column=COL["cum_pnl"],
                    value=_row_formulas(rr)["cum_pnl"])

    wb.save(wb_path)
    wb.close()
    return len(new_trades)


# --- Analysis write-back ---------------------------------------------------
_HDR_FILL = PatternFill("solid", fgColor="1F2937")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_SEC_FONT = Font(bold=True, size=13, color="111827")


def write_analysis_sheets(metrics: dict, ai_text: str,
                          workbook: Path | None = None) -> None:
    """Write/replace 'Analysis' and 'AI Suggestions' sheets in the workbook."""
    wb_path = Path(workbook) if workbook else config.WORKBOOK
    backup(wb_path)
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    for name in ("Analysis", "AI Suggestions"):
        if name in wb.sheetnames:
            del wb[name]

    a = wb.create_sheet("Analysis")
    a.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        a.column_dimensions[col].width = 14
    row = 1

    def section(title):
        nonlocal row
        c = a.cell(row=row, column=1, value=title)
        c.font = _SEC_FONT
        row += 1

    def header(cols):
        nonlocal row
        for i, h in enumerate(cols, start=1):
            c = a.cell(row=row, column=i, value=h)
            c.fill, c.font = _HDR_FILL, _HDR_FONT
        row += 1

    def line(cols):
        nonlocal row
        for i, v in enumerate(cols, start=1):
            a.cell(row=row, column=i, value=v)
        row += 1

    s = metrics["summary"]
    section("Overview")
    line(["Generated", metrics.get("generated_at", "")])
    line(["Date range", metrics.get("date_range", "")])
    line(["Trades", s["n_trades"]])
    line(["Win rate", f'{s["win_rate"]*100:.1f}%'])
    line(["Total P&L", round(s["total_pnl"], 2)])
    line(["Profit factor", s["profit_factor"]])
    line(["Expectancy / trade", round(s["expectancy"], 2)])
    line(["Avg win", round(s["avg_win"], 2)])
    line(["Avg loss", round(s["avg_loss"], 2)])
    line(["Avg RR targeted", s["avg_rr_targeted"]])
    line(["Avg RR realized", s["avg_rr_realized"]])
    line(["Max drawdown", round(s["max_drawdown"], 2)])
    line(["Max win streak", s["max_win_streak"]])
    line(["Max loss streak", s["max_loss_streak"]])
    row += 1

    def table(title, key, label):
        nonlocal row
        rows = metrics.get(key, [])
        if not rows:
            return
        section(title)
        header([label, "Trades", "Win %", "Total P&L", "Expectancy"])
        for b in rows:
            line([b["name"], b["n"], f'{b["win_rate"]*100:.0f}%',
                  round(b["total_pnl"], 2), round(b["expectancy"], 2)])
        row += 1

    table("By Setup", "by_setup", "Setup")
    table("By Instrument", "by_instrument", "Instrument")
    table("By Direction", "by_direction", "Direction")
    table("By Hour", "by_hour", "Hour")
    table("By Day of Week", "by_dayofweek", "Day")

    flags = metrics.get("behavior_flags", [])
    if flags:
        section("Behavior Flags")
        for f in flags:
            line([f["type"], f["detail"]])

    # AI Suggestions sheet
    ai = wb.create_sheet("AI Suggestions")
    ai.column_dimensions["A"].width = 120
    ai.cell(row=1, column=1, value="Claude AI Coaching").font = _SEC_FONT
    rr = 3
    for para in (ai_text or "No AI output.").split("\n"):
        c = ai.cell(row=rr, column=1, value=para)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1

    wb.save(wb_path)
    wb.close()
