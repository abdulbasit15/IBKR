"""Entry point for the intraday long-only equity bots (PyInstaller one-file target).

Loads equity.json (next to sys.executable when frozen), takes ONE start-of-day equity
snapshot + detects STK volume scale on a bootstrap connection, builds the shared
PortfolioRiskManager + rate limiter + day cache + journal, then launches one thread per
active strategy (each with its own clientId and event loop).

Run:  python runner.py            (or the frozen .exe)
TWS/IB Gateway must be running on the configured port with the API enabled.
Defaults to PAPER account DU672616.
"""
from __future__ import annotations
import asyncio
import json
import os
import sys
import threading
import time

from ib_async import IB


def base_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


BASE = base_dir()
if BASE not in sys.path:
    sys.path.insert(0, BASE)

import calendar_util as cal                     # noqa: E402
from market_data import RateLimiter, DailyCache, detect_volume_scale  # noqa: E402
from portfolio_risk import PortfolioRiskManager, SymbolLock  # noqa: E402
from reporting import TradeReporter  # noqa: E402
from strategies.orb_stocks_in_play import ORBStocksInPlay  # noqa: E402
from strategies.nr7_compression import NR7Compression      # noqa: E402
from strategies.pdh_breakout import PDHBreakout            # noqa: E402

REGISTRY = {
    "orb_stocks_in_play": ORBStocksInPlay,
    "nr7_compression": NR7Compression,
    "pdh_breakout": PDHBreakout,
}

JOURNAL_HEADERS = ["Event", "Symbol", "Sector", "Strategy", "Shares", "Entry", "Stop",
                   "Target", "RiskAtStop", "Exit", "PnL", "R_Multiple", "Result"]
_journal_lock = threading.Lock()
_log_lock = threading.Lock()


def make_logger(log_path: str):
    def log(msg: str):
        line = f"[{cal.now_et().strftime('%Y-%m-%d %H:%M:%S')} ET] {msg}"
        with _log_lock:
            print(line, flush=True)
            try:
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(line + "\n")
            except OSError:
                pass
    return log


def make_journal(path: str):
    def journal(strategy: str, row: dict):
        with _journal_lock:
            try:
                import openpyxl
                if os.path.exists(path):
                    wb = openpyxl.load_workbook(path)
                else:
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                sheet = (strategy or "trades")[:31]
                if sheet not in wb.sheetnames:
                    ws = wb.create_sheet(sheet)
                    ws.append(["Time"] + JOURNAL_HEADERS)
                else:
                    ws = wb[sheet]
                ws.append([cal.now_et().strftime("%Y-%m-%d %H:%M:%S")] +
                          [row.get(h, "") for h in JOURNAL_HEADERS])
                wb.save(path)
            except Exception as e:  # journaling must never crash trading
                print(f"[journal error] {e}", flush=True)
    return journal


def bootstrap(cfg, log):
    """Main-thread bootstrap connection: NetLiquidation snapshot + volume-scale detect."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    ib = IB()
    account = cfg.get("default_account", "")
    equity, scale = 0.0, 1
    try:
        ib.connect(cfg.get("host", "127.0.0.1"), int(cfg.get("port", 7497)),
                   clientId=int(cfg.get("client_id_base", 30)) + 90, account=account)
        try:
            ib.reqMarketDataType(int(cfg.get("market_data_type", 1)))
        except Exception:
            pass
        for v in ib.accountValues(account):
            if v.tag == "NetLiquidation" and (not v.currency or v.currency == "USD"):
                equity = float(v.value)
                break
        scale = detect_volume_scale(ib)
        log(f"bootstrap: NetLiquidation={equity} volume_scale={scale}")
    except Exception as e:
        log(f"bootstrap FAILED ({e}); cannot size positions without equity. Aborting.")
    finally:
        try:
            ib.disconnect()
        except Exception:
            pass
    return equity, scale


def main():
    cfg_path = os.path.join(BASE, "equity.json")
    with open(cfg_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    log_dir = os.path.join(BASE, "logs")
    os.makedirs(log_dir, exist_ok=True)
    stamp = cal.now_et().strftime("%Y%m%d")
    log = make_logger(os.path.join(log_dir, f"equity_{stamp}.log"))
    journal = make_journal(os.path.join(BASE, f"equity_journal_{stamp}.xlsx"))

    if not cal.is_trading_day():
        log("Not a trading day (holiday/weekend). Exiting.")
        return

    equity, vol_scale = bootstrap(cfg, log)
    if equity <= 0:
        log("No equity snapshot; aborting (check TWS connection / account).")
        return

    shared_risk = cfg.get("shared_risk", {})
    reports_dir = os.path.join(BASE, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    shared = {
        "host": cfg.get("host", "127.0.0.1"),
        "port": int(cfg.get("port", 7497)),
        "default_account": cfg.get("default_account"),
        "market_data_type": int(cfg.get("market_data_type", 1)),  # 1=live 2=frozen 3=delayed 4=delayed-frozen
        "shared_risk": shared_risk,
        "sector_map": cfg.get("sector_map", {}),
        "rate_limiter": RateLimiter(min_interval=float(cfg.get("hist_min_interval_sec", 2.0))),
        "cache": DailyCache(os.path.join(BASE, f"cache_{stamp}.json"), stamp),
        "vol_scale": vol_scale,
        "start_equity": equity,
        "journal": journal,
    }
    symbol_lock = SymbolLock()   # cross-strategy: never two strategies long the same symbol

    active = cfg.get("active_strategies", [])
    base_id = int(cfg.get("client_id_base", 30))
    overrides = ("max_concurrent_tickers", "max_positions_per_sector", "daily_loss_limit_pct",
                 "aggregate_open_risk_pct", "risk_per_trade_pct")
    threads, managers = [], {}
    for i, name in enumerate(active):
        block = cfg.get("strategies", {}).get(name)
        if not block:
            log(f"active strategy '{name}' has no config block; skipping")
            continue
        cls = REGISTRY.get(block.get("strategy_type"))
        if not cls:
            log(f"unknown strategy_type for '{name}'; skipping")
            continue
        block.setdefault("client_id", base_id + i)
        safe = "".join(ch if ch.isalnum() else "_" for ch in name)
        # per-strategy daily log + persistent analytics report + own risk book
        slog = make_logger(os.path.join(log_dir, f"{safe}_{stamp}.log"))
        reporter = TradeReporter(os.path.join(reports_dir, f"report_{safe}.xlsx"), name)
        capital = float(block.get("strategy_capital", equity))
        risk_cfg = dict(shared_risk)
        risk_cfg.update({k: block[k] for k in overrides if k in block})
        rm = PortfolioRiskManager(capital, risk_cfg, symbol_lock=symbol_lock,
                                  state_path=os.path.join(BASE, f"risk_{safe}_{stamp}.json"))
        managers[name] = rm
        inst = cls(name, block, shared, rm, slog, reporter=reporter)
        t = threading.Thread(target=inst.run, name=name, daemon=False)
        threads.append(t)
        t.start()
        mt = risk_cfg.get("max_concurrent_tickers", risk_cfg.get("max_concurrent_positions", 5))
        log(f"launched '{name}' ({block['strategy_type']}) clientId={block['client_id']} "
            f"capital={capital:,.0f} fixed_stocks={block.get('fixed_stocks', 0)} max_tickers={mt}")
        time.sleep(1.0)  # stagger connections

    for t in threads:
        t.join()
    for name, rm in managers.items():
        log(f"[{name}] final risk snapshot: {rm.snapshot()}")


if __name__ == "__main__":
    main()
