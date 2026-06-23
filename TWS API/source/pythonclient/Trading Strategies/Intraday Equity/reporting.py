"""Per-strategy trade report.

record_trade() appends one CLOSED trade and then recomputes the Daily, ByTicker (per
day per ticker), and Summary sheets from the full Trades history. The workbook is NOT
date-stamped, so it ACCUMULATES across days and always reflects the latest info.

Sheets:
  Trades    - every closed trade (one row each), appended forever
  Daily     - per trading day: trades, wins, losses, win-rate, gross P/L, avg R
  ByTicker  - per (day, ticker): trades, wins, win-rate, gross P/L, avg R
  Summary   - overall strategy: totals, win-rate, P/L, avg R, profit factor, best/worst
"""
from __future__ import annotations
import os
import threading
from collections import defaultdict

COLS = ["Date", "Time", "Strategy", "Ticker", "Sector", "Shares", "Entry", "Stop",
        "Target", "Exit", "PnL", "R_Multiple", "Result", "Reason", "HoldMin"]


def _f(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


class TradeReporter:
    def __init__(self, path: str, strategy_name: str):
        self.path = path
        self.name = strategy_name
        self._lock = threading.Lock()

    def record_trade(self, row: dict) -> None:
        with self._lock:
            try:
                import openpyxl
                if os.path.exists(self.path):
                    wb = openpyxl.load_workbook(self.path)
                else:
                    wb = openpyxl.Workbook()
                # Trades sheet (reuse the default 'Sheet' if pristine)
                if "Trades" not in wb.sheetnames:
                    default = wb.active
                    if default is not None and default.title == "Sheet" and default.max_row <= 1:
                        ws = default
                        ws.title = "Trades"
                    else:
                        ws = wb.create_sheet("Trades")
                    ws.append(COLS)
                else:
                    ws = wb["Trades"]
                ws.append([row.get(c, "") for c in COLS])

                trades = [dict(zip(COLS, [c.value for c in r]))
                          for r in ws.iter_rows(min_row=2)]
                self._rebuild_daily(wb, trades)
                self._rebuild_by_ticker(wb, trades)
                self._rebuild_summary(wb, trades)
                wb.save(self.path)
            except Exception as e:  # reporting must never crash trading
                print(f"[reporter error] {e}", flush=True)

    @staticmethod
    def _fresh(wb, name):
        if name in wb.sheetnames:
            del wb[name]
        return wb.create_sheet(name)

    def _rebuild_daily(self, wb, trades):
        ws = self._fresh(wb, "Daily")
        ws.append(["Date", "Trades", "Wins", "Losses", "WinRate%", "GrossPnL", "AvgR"])
        g = defaultdict(list)
        for t in trades:
            g[t.get("Date", "")].append(t)
        for d in sorted(k for k in g if k):
            ts = g[d]
            n = len(ts)
            wins = sum(1 for t in ts if _f(t.get("PnL")) > 0)
            losses = sum(1 for t in ts if _f(t.get("PnL")) < 0)
            pnl = sum(_f(t.get("PnL")) for t in ts)
            avgr = sum(_f(t.get("R_Multiple")) for t in ts) / n if n else 0
            ws.append([d, n, wins, losses, round(100 * wins / n, 1) if n else 0,
                       round(pnl, 2), round(avgr, 3)])

    def _rebuild_by_ticker(self, wb, trades):
        ws = self._fresh(wb, "ByTicker")
        ws.append(["Date", "Ticker", "Trades", "Wins", "WinRate%", "GrossPnL", "AvgR"])
        g = defaultdict(list)
        for t in trades:
            g[(t.get("Date", ""), t.get("Ticker", ""))].append(t)
        for (d, tk) in sorted(k for k in g if k[0]):
            ts = g[(d, tk)]
            n = len(ts)
            wins = sum(1 for t in ts if _f(t.get("PnL")) > 0)
            pnl = sum(_f(t.get("PnL")) for t in ts)
            avgr = sum(_f(t.get("R_Multiple")) for t in ts) / n if n else 0
            ws.append([d, tk, n, wins, round(100 * wins / n, 1) if n else 0,
                       round(pnl, 2), round(avgr, 3)])

    def _rebuild_summary(self, wb, trades):
        ws = self._fresh(wb, "Summary")
        n = len(trades)
        wins = [t for t in trades if _f(t.get("PnL")) > 0]
        losses = [t for t in trades if _f(t.get("PnL")) < 0]
        pnl = sum(_f(t.get("PnL")) for t in trades)
        gross_win = sum(_f(t.get("PnL")) for t in wins)
        gross_loss = abs(sum(_f(t.get("PnL")) for t in losses))
        pf = round(gross_win / gross_loss, 2) if gross_loss > 0 else ("inf" if gross_win > 0 else 0)
        avgr = round(sum(_f(t.get("R_Multiple")) for t in trades) / n, 3) if n else 0
        largest_win = round(max((_f(t.get("PnL")) for t in trades), default=0), 2)
        largest_loss = round(min((_f(t.get("PnL")) for t in trades), default=0), 2)
        days = len({t.get("Date", "") for t in trades if t.get("Date")})
        rows = [
            ("Strategy", self.name),
            ("Trading days", days),
            ("Total trades", n),
            ("Wins", len(wins)),
            ("Losses", len(losses)),
            ("Win rate %", round(100 * len(wins) / n, 1) if n else 0),
            ("Total P/L", round(pnl, 2)),
            ("Avg P/L per trade", round(pnl / n, 2) if n else 0),
            ("Avg R multiple", avgr),
            ("Profit factor", pf),
            ("Largest win", largest_win),
            ("Largest loss", largest_loss),
        ]
        ws.append(["Metric", "Value"])
        for k, v in rows:
            ws.append([k, v])
